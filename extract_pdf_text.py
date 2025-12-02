#!/usr/bin/env python3
"""
PDF to Excel Converter for Compressor Data

Extracts compressor data from PDF files and converts directly to Excel format.
Supports single file, multiple files, and directory-based batch processing.

Features:
- Extracts Date and Consumo values from PDF pages
- Automatically identifies compressor number from filename
- Writes data to Excel with Date, Consumo, and Compressor columns
- Memory-efficient processing for large PDFs (6000+ pages)
- Batch processing for multiple PDFs with error recovery

Usage:
    Single file:  python extract_pdf_text.py input.pdf
    Multiple:     python extract_pdf_text.py file1.pdf file2.pdf -o output.xlsx
    Directory:    python extract_pdf_text.py -d folder/ -o output.xlsx

Libraries:
- PyMuPDF (fitz) - Fast extraction with good text quality (active)
- pdfplumber - Excellent for table extraction (disabled)
- pdfminer.six - Low-level control for complex layouts (disabled)
"""

import time
import os
import sys
import re
from pathlib import Path
from datetime import datetime


def _generate_fallback_compressor_name(filename_only):
    """
    Generate a fallback compressor name from filename when pattern matching fails.
    
    Args:
        filename_only: Just the filename (no path)
    
    Returns:
        str: Fallback compressor name
    """
    # Try to extract any number from the filename
    # Look for patterns like: COMPRESOR1, COMPRESOR-1, compresor1, etc.
    number_patterns = [
        r'(?i)compresor[-_]?(\d+)',  # COMPRESOR1, COMPRESOR-1, COMPRESOR_1
        r'(\d+)',  # Any number in the filename
    ]
    
    for pattern in number_patterns:
        match = re.search(pattern, filename_only)
        if match:
            compressor_num = match.group(1)
            return f"Compressor {compressor_num}"
    
    # If no number found, use filename stem (without extension) as fallback
    filename_stem = Path(filename_only).stem
    # Clean up the stem - remove common separators and use first meaningful part
    cleaned_stem = filename_stem.split('-')[0].split('_')[0]
    if cleaned_stem:
        return f"Compressor ({cleaned_stem})"
    
    # Last resort: generic name
    return "Compressor (Unknown)"


def extract_compressor_from_pdf_path(pdf_path, fallback=True):
    """
    Extract compressor number from PDF filename.
    
    Examples:
        COMPRESOR4-ABR-JUN-25.PDF -> "Compressor 4"
        compresor1-test.pdf -> "Compressor 1"
        COMPRESOR12-data.pdf -> "Compressor 12"
        COMPRESOR1-ABR-MAY-25.PDF -> "Compressor 1"
        COMPRESOR4-ENE-MAR-25.PDF -> "Compressor 4"
        COMPRESOR5-JUL-SEPT-25.PDF -> "Compressor 5"
    
    Args:
        pdf_path: Path object or string to PDF file
        fallback: If True, use fallback name generation when pattern matching fails.
                 If False, raise ValueError when pattern doesn't match.
    
    Returns:
        str: Formatted compressor name like "Compressor 4"
    
    Raises:
        ValueError: If compressor number cannot be found and fallback=False
    """
    # Extract just the filename if full path is provided
    filename_only = Path(pdf_path).name
    
    # Try multiple patterns to find COMPRESOR followed by digits
    patterns = [
        r'(?i)compresor(\d+)',      # COMPRESOR4, compresor1
        r'(?i)compresor[-_](\d+)',  # COMPRESOR-4, COMPRESOR_4
        r'(?i)compresor\s*(\d+)',   # COMPRESOR 4 (with space)
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename_only)
        if match:
            compressor_num = match.group(1)
            return f"Compressor {compressor_num}"
    
    # If no pattern matched and fallback is enabled, use fallback
    if fallback:
        fallback_name = _generate_fallback_compressor_name(filename_only)
        return fallback_name
    
    # If fallback is disabled, raise error
    raise ValueError(
        f"Could not extract compressor number from filename: {filename_only}. "
        f"Expected pattern: COMPRESOR<number>"
    )


def parse_page_text(page_text):
    """
    Parse page text and extract Date and Consumo values.
    
    Skips page markers, headers, and empty lines. Extracts data rows
    matching pattern: DD/MM/YYYY H:MM:SS VALUE or DD/MM/YYYY HH:MM:SS VALUE
    
    Args:
        page_text: Text content from a single PDF page
    
    Returns:
        list of tuples: [(date_str, consumo_value), ...]
    """
    if not page_text:
        return []
    
    # Pattern to match: DD/MM/YYYY H:MM:SS VALUE or DD/MM/YYYY HH:MM:SS VALUE
    # Date pattern: \d{2}/\d{2}/\d{4}
    # Time pattern: \d{1,2}:\d{2}:\d{2}  (handles both single and double digit hours)
    # Value pattern: \d+ (integer) or \d+\.\d+ (float)
    data_pattern = re.compile(r'^(\d{2}/\d{2}/\d{4} \d{1,2}:\d{2}:\d{2})\s+([\d.]+)$')
    
    data_rows = []
    lines = page_text.split('\n')
    
    for line in lines:
        line = line.strip()
        
        # Skip empty lines, page markers, and header lines
        if not line or line.startswith('===') or 'Página' in line:
            continue
        
        # Skip header text
        if any(header in line for header in ['HISTORICO', 'Fecha', 'Hora', 'Consumo motor', 'compresor']):
            continue
        
        # Try to match data pattern
        match = data_pattern.match(line)
        if match:
            date_str = match.group(1)
            consumo_str = match.group(2)
            
            # Convert consumo to numeric (handle both int and float)
            try:
                consumo = float(consumo_str) if '.' in consumo_str else int(consumo_str)
            except ValueError:
                # Skip malformed consumo values
                continue
            
            data_rows.append((date_str, consumo))
    
    return data_rows


def extract_with_pdfplumber(pdf_path, output_path):
    """Extract text using pdfplumber library."""
    import pdfplumber
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Extracting with pdfplumber...")
    print(f"  Opening PDF file...")
    start_time = time.time()
    open_start = time.time()
    
    page_count = 0
    
    try:
        BATCH_SIZE = 100
        
        with pdfplumber.open(pdf_path) as pdf:
            open_time = time.time() - open_start
            print(f"  ✓ PDF opened in {open_time:.2f} seconds")
            
            # Get page count from PDF metadata without triggering page cache
            # Access doc.num_pages if available, otherwise detect dynamically
            try:
                page_count = pdf.doc.num_pages if hasattr(pdf.doc, 'num_pages') else None
            except:
                page_count = None
            
            if page_count is None:
                # Fallback: try to get from PDF catalog, or we'll detect dynamically
                print(f"  Detecting page count dynamically...")
                page_count = "?"  # Will be determined during processing
            
            print(f"  Starting page extraction (batching {BATCH_SIZE} pages at a time, flushing cache after each page)...")
            
            extract_start = time.time()
            last_log_time = extract_start
            
            # Incremental statistics instead of storing all page_times
            total_time = 0.0
            page_count_actual = 0
            min_time = float('inf')
            max_time = 0.0
            slowest_page_num = 0
            
            batch_buffer = []  # Buffer to collect pages before writing
            file_size_bytes = 0  # Track file size manually
            
            # Open file once and keep it open between batches
            output_file = open(output_path, 'w', encoding='utf-8')
            
            i = 0
            CACHE_CLEAR_INTERVAL = 500  # Clear pdf._pages cache every 500 pages (more frequent)
            
            while True:
                try:
                    # Access page directly by index
                    # NOTE: pdf.pages[i] triggers the pages property which caches ALL pages
                    # We'll clear the cache periodically to prevent accumulation
                    page = pdf.pages[i]
                    page_start = time.time()
                    
                    page_text = page.extract_text()
                    page_time = time.time() - page_start
                    
                    # Update incremental statistics
                    page_count_actual = i + 1
                    total_time += page_time
                    if page_time < min_time:
                        min_time = page_time
                    if page_time > max_time:
                        max_time = page_time
                        slowest_page_num = i + 1
                    
                    # Add page to batch buffer
                    if page_text:
                        batch_buffer.append(f"=== PAGE {i + 1} ===\n{page_text}\n")
                    
                    # Flush page cache to free memory immediately
                    try:
                        page.flush_cache()
                    except AttributeError:
                        # If flush_cache doesn't exist, try close()
                        try:
                            page.close()
                        except:
                            pass
                    
                    # Write batch every BATCH_SIZE pages
                    if len(batch_buffer) >= BATCH_SIZE:
                        batch_text = ''.join(batch_buffer)
                        output_file.write(batch_text)
                        output_file.flush()
                        file_size_bytes += len(batch_text.encode('utf-8'))
                        batch_buffer = []  # Clear buffer to free memory
                        
                        # Clear pdf._pages cache after each batch write to prevent accumulation
                        # This is critical: pdf.pages[i] caches ALL pages, so we must clear regularly
                        if hasattr(pdf, '_pages') and pdf._pages:
                            # Only clear if we have more than BATCH_SIZE pages cached
                            # This prevents clearing too early when cache is small
                            if len(pdf._pages) > BATCH_SIZE:
                                pdf._pages = []
                                # Force garbage collection hint
                                import gc
                                gc.collect()
                    
                    # Also clear cache at regular intervals as backup
                    if i > 0 and i % CACHE_CLEAR_INTERVAL == 0:
                        if hasattr(pdf, '_pages') and pdf._pages:
                            pdf._pages = []
                            import gc
                            gc.collect()
                            print(f"  [Cache cleared at page {i}]")
                    
                    # Log progress every 10 pages or every 5 seconds, whichever comes first
                    current_time = time.time()
                    if (i + 1) % 10 == 0 or (current_time - last_log_time) >= 5:
                        elapsed = current_time - extract_start
                        avg_time_per_page = elapsed / (i + 1)
                        remaining_pages = (page_count_actual - (i + 1)) if isinstance(page_count, int) else None
                        est_remaining = avg_time_per_page * remaining_pages if remaining_pages else None
                        file_size_mb = file_size_bytes / (1024 * 1024)
                        batch_num = (i + 1) // BATCH_SIZE + (1 if (i + 1) % BATCH_SIZE > 0 else 0)
                        page_display = f"{i + 1}/{page_count}" if isinstance(page_count, int) else f"{i + 1} (detecting...)"
                        est_str = f"{est_remaining/60:.1f} min" if est_remaining else "calculating..."
                        print(f"  [{datetime.now().strftime('%H:%M:%S')}] Page {page_display} "
                              f"(batch {batch_num}, "
                              f"avg: {avg_time_per_page:.3f}s/page, "
                              f"last: {page_time:.3f}s, "
                              f"file: {file_size_mb:.2f} MB, "
                              f"est. remaining: {est_str})")
                        last_log_time = current_time
                        sys.stdout.flush()
                    
                    i += 1
                    
                except (IndexError, AttributeError) as e:
                    # Reached end of pages or page access failed
                    break
            
            # Write remaining batch
            if batch_buffer:
                batch_text = ''.join(batch_buffer)
                output_file.write(batch_text)
                file_size_bytes += len(batch_text.encode('utf-8'))
            
            output_file.close()
            
            page_count = page_count_actual
            
            extract_time = time.time() - extract_start
            avg_time = total_time / page_count if page_count > 0 else 0
            print(f"  ✓ Page extraction completed in {extract_time:.2f} seconds")
            print(f"  Average time per page: {avg_time:.3f} seconds")
            print(f"  Fastest page: {min_time:.3f} seconds")
            print(f"  Slowest page: {max_time:.3f} seconds (page {slowest_page_num})")
            
            # Read file to get character count
            print(f"  Reading file for statistics...")
            with open(output_path, 'r', encoding='utf-8') as f:
                full_text = f.read()
            
    except Exception as e:
        print(f"  ✗ Error with pdfplumber: {e}")
        import traceback
        traceback.print_exc()
        if 'output_file' in locals():
            try:
                output_file.close()
            except:
                pass
        return None, None, None
    
    elapsed_time = time.time() - start_time
    
    char_count = len(full_text)
    file_size_mb = file_size_bytes / (1024 * 1024)
    print(f"  ✓ File complete: {file_size_mb:.2f} MB, {char_count:,} characters")
    
    return page_count, char_count, elapsed_time


def _extract_pdf_pages(pdf_path, compressor_name, worksheet, progress_callback=None):
    """
    Core PDF extraction logic that processes pages and writes to worksheet.
    
    This is the shared extraction logic used by both single-file and multi-file modes.
    
    Args:
        pdf_path: Path to PDF file
        compressor_name: Compressor name string (e.g., "Compressor 4")
        worksheet: openpyxl WriteOnlyWorksheet to append rows to
        progress_callback: Optional callback function(page_num, total_pages, rows_written)
    
    Returns:
        tuple: (page_count, rows_written, stats_dict)
        stats_dict contains: total_time, min_time, max_time, slowest_page_num
    """
    import fitz  # PyMuPDF
    
    BATCH_SIZE = 100
    
    doc = fitz.open(pdf_path)
    page_count = len(doc)
    
    extract_start = time.time()
    
    # Incremental statistics
    total_time = 0.0
    min_time = float('inf')
    max_time = 0.0
    slowest_page_num = 0
    
    # Buffer to collect parsed rows before writing to Excel
    excel_batch_buffer = []
    total_rows_written = 0
    
    for i in range(page_count):
        page_start = time.time()
        
        page = doc[i]
        page_text = page.get_text()
        page_time = time.time() - page_start
        
        # Update incremental statistics
        total_time += page_time
        if page_time < min_time:
            min_time = page_time
        if page_time > max_time:
            max_time = page_time
            slowest_page_num = i + 1
        
        # Parse page text immediately to extract Date/Consumo values
        if page_text:
            parsed_rows = parse_page_text(page_text)
            
            # Add compressor name to each row and add to batch buffer
            for date_str, consumo in parsed_rows:
                excel_batch_buffer.append([date_str, consumo, compressor_name])
        
        # Write batch to Excel every BATCH_SIZE pages
        if len(excel_batch_buffer) >= BATCH_SIZE:
            # Write all rows in batch to Excel
            for row in excel_batch_buffer:
                worksheet.append(row)
            total_rows_written += len(excel_batch_buffer)
            excel_batch_buffer = []  # Clear buffer to free memory
            
            # Call progress callback if provided
            if progress_callback:
                progress_callback(i + 1, page_count, total_rows_written)
    
    # Write remaining batch to Excel
    if excel_batch_buffer:
        for row in excel_batch_buffer:
            worksheet.append(row)
        total_rows_written += len(excel_batch_buffer)
        excel_batch_buffer = []
    
    doc.close()
    
    stats = {
        'total_time': total_time,
        'min_time': min_time,
        'max_time': max_time,
        'slowest_page_num': slowest_page_num,
        'extract_time': time.time() - extract_start
    }
    
    return page_count, total_rows_written, stats


def extract_pdf_to_worksheet(pdf_path, worksheet, compressor_name=None, show_progress=True):
    """
    Extract data from a PDF and append rows to an existing Excel worksheet.
    
    Used for multi-file processing where multiple PDFs write to the same worksheet.
    
    Args:
        pdf_path: Path to PDF file
        worksheet: openpyxl WriteOnlyWorksheet to append rows to
        compressor_name: Optional compressor name (extracted from filename if not provided)
        show_progress: Whether to show progress messages
    
    Returns:
        tuple: (success: bool, page_count: int, rows_written: int, error_message: str or None)
    """
    try:
        # Extract compressor number from PDF filename if not provided
        # Always use fallback=True to ensure PDFs are processed even if name extraction fails
        if compressor_name is None:
            compressor_name = extract_compressor_from_pdf_path(pdf_path, fallback=True)
        
        if show_progress:
            print(f"  Processing: {Path(pdf_path).name}")
            print(f"    Compressor: {compressor_name}")
        
        # Progress callback for detailed logging
        def progress_callback(page_num, total_pages, rows_written):
            if show_progress and (page_num % 10 == 0 or page_num == total_pages):
                print(f"    Page {page_num}/{total_pages}, rows extracted: {rows_written:,}")
        
        page_count, rows_written, stats = _extract_pdf_pages(
            pdf_path, compressor_name, worksheet, progress_callback
        )
        
        if show_progress:
            print(f"    ✓ Completed: {page_count} pages, {rows_written:,} rows extracted")
        
        return True, page_count, rows_written, None
        
    except Exception as e:
        error_msg = str(e)
        if show_progress:
            print(f"    ✗ Error: {error_msg}")
        return False, 0, 0, error_msg


def extract_with_pymupdf(pdf_path, output_path):
    """Extract text using PyMuPDF (fitz) library and write directly to Excel (single-file mode)."""
    from openpyxl import Workbook
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Extracting with PyMuPDF (fitz) and writing to Excel...")
    print(f"  Opening PDF file...")
    start_time = time.time()
    
    workbook = None
    worksheet = None
    
    try:
        # Extract compressor number from PDF filename
        # Use fallback=True to ensure processing even if name extraction fails
        compressor_name = extract_compressor_from_pdf_path(pdf_path, fallback=True)
        print(f"  Extracted compressor: {compressor_name}")
        
        # Initialize Excel workbook in write-only mode (memory-efficient)
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(title="Compressor Data")
        
        # Write headers as the first row: Date, Consumo, Compressor
        worksheet.append(['Date', 'Consumo', 'Compressor'])
        print(f"  ✓ Excel headers written: Date, Consumo, Compressor")
        
        # Use shared extraction logic
        last_log_time = time.time()
        
        def progress_callback(page_num, total_pages, rows_written):
            nonlocal last_log_time
            current_time = time.time()
            if page_num % 10 == 0 or (current_time - last_log_time) >= 5:
                elapsed = current_time - start_time
                avg_time_per_page = elapsed / page_num if page_num > 0 else 0
                remaining_pages = total_pages - page_num
                est_remaining = avg_time_per_page * remaining_pages
                batch_num = page_num // 100 + (1 if page_num % 100 > 0 else 0)
                print(f"  [{datetime.now().strftime('%H:%M:%S')}] Page {page_num}/{total_pages} "
                      f"(batch {batch_num}, "
                      f"avg: {avg_time_per_page:.3f}s/page, "
                      f"rows: {rows_written:,}, "
                      f"est. remaining: {est_remaining/60:.1f} min)")
                last_log_time = current_time
                sys.stdout.flush()
        
        page_count, total_rows_written, stats = _extract_pdf_pages(
            pdf_path, compressor_name, worksheet, progress_callback
        )
        
        # Save Excel workbook
        print(f"  Saving Excel file...")
        workbook.save(output_path)
        workbook.close()
        
        extract_time = stats['extract_time']
        avg_time = stats['total_time'] / page_count if page_count > 0 else 0
        print(f"  ✓ Page extraction completed in {extract_time:.2f} seconds")
        print(f"  Average time per page: {avg_time:.3f} seconds")
        print(f"  Fastest page: {stats['min_time']:.3f} seconds")
        print(f"  Slowest page: {stats['max_time']:.3f} seconds (page {stats['slowest_page_num']})")
        
        # Get file size
        file_size_bytes = Path(output_path).stat().st_size
        file_size_mb = file_size_bytes / (1024 * 1024)
        print(f"  ✓ Excel file complete: {file_size_mb:.2f} MB, {total_rows_written:,} data rows")
        
    except Exception as e:
        print(f"  ✗ Error with PyMuPDF/Excel: {e}")
        import traceback
        traceback.print_exc()
        # Try to save workbook if it exists
        if workbook is not None:
            try:
                workbook.save(output_path)
                workbook.close()
                print(f"  ✓ Saved partial Excel file: {output_path}")
            except:
                pass
        return None, None, None
    
    elapsed_time = time.time() - start_time
    
    return page_count, total_rows_written, elapsed_time


def sanitize_excel_sheet_name(name, max_length=31):
    """
    Sanitize a compressor name for use as an Excel sheet name.
    
    Excel sheet name limitations:
    - Maximum 31 characters
    - Cannot contain: : \ / ? * [ ]
    - Cannot be empty
    
    Args:
        name: Compressor name string
        max_length: Maximum length for sheet name (default: 31)
    
    Returns:
        str: Sanitized sheet name safe for Excel
    """
    # Replace invalid characters with underscore
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    
    # Truncate if too long
    if len(sanitized) > max_length:
        sanitized = sanitized[:max_length]
    
    # Ensure not empty
    if not sanitized.strip():
        sanitized = "Compressor"
    
    return sanitized


def find_pdf_files(directory):
    """
    Find all PDF files in a directory.
    
    Args:
        directory: Path to directory (string or Path object)
    
    Returns:
        list: Sorted list of Path objects for all PDF files found
    """
    directory = Path(directory)
    
    if not directory.exists():
        raise FileNotFoundError(f"Directory not found: {directory}")
    
    if not directory.is_dir():
        raise ValueError(f"Path is not a directory: {directory}")
    
    # Find all PDF files (case-insensitive)
    pdf_files = sorted(directory.glob("*.pdf")) + sorted(directory.glob("*.PDF"))
    
    # Remove duplicates (in case both patterns matched same files)
    pdf_files = list(dict.fromkeys(pdf_files))
    
    return pdf_files


def process_multiple_pdfs(pdf_files, excel_output_path):
    """
    Process multiple PDF files and combine all data into a single Excel file.
    Creates separate sheets for each compressor to handle Excel's row limit (1,048,576 rows per sheet).
    
    Args:
        pdf_files: List of PDF file paths (Path objects or strings)
        excel_output_path: Path where Excel file should be written
    
    Returns:
        dict: Summary with keys: success_count, failure_count, total_rows, 
              failed_files (list of tuples: (file_path, error_message)),
              compressor_stats (dict mapping compressor_name -> row_count)
    """
    from openpyxl import Workbook
    
    if not pdf_files:
        raise ValueError("No PDF files provided")
    
    # Validate all PDF files exist
    for pdf_file in pdf_files:
        pdf_path = Path(pdf_file)
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Processing {len(pdf_files)} PDF file(s)...")
    print(f"  Output file: {excel_output_path}")
    print()
    
    workbook = None
    worksheets = {}  # Dictionary mapping compressor_name -> worksheet
    compressor_stats = {}  # Dictionary mapping compressor_name -> row_count
    
    success_count = 0
    failure_count = 0
    total_rows = 0
    failed_files = []
    
    overall_start_time = time.time()
    
    try:
        # Initialize Excel workbook in write-only mode (memory-efficient)
        workbook = Workbook(write_only=True)
        # In write-only mode, we can't access worksheets list, so we'll just create sheets as needed
        # The default sheet won't be used
        
        # Process each PDF file
        for idx, pdf_file in enumerate(pdf_files, 1):
            pdf_path = Path(pdf_file)
            print(f"[{idx}/{len(pdf_files)}] Processing: {pdf_path.name}")
            
            try:
                # Extract compressor name first to determine which sheet to use
                compressor_name = extract_compressor_from_pdf_path(pdf_path, fallback=True)
                
                # Check if worksheet exists for this compressor
                if compressor_name not in worksheets:
                    # Create new worksheet for this compressor
                    sheet_title = sanitize_excel_sheet_name(compressor_name)
                    worksheet = workbook.create_sheet(title=sheet_title)
                    
                    # Write headers as the first row: Date, Consumo, Compressor
                    worksheet.append(['Date', 'Consumo', 'Compressor'])
                    worksheets[compressor_name] = worksheet
                    compressor_stats[compressor_name] = 0
                    print(f"  Created sheet: {sheet_title}")
                else:
                    # Use existing worksheet for this compressor
                    worksheet = worksheets[compressor_name]
                    print(f"  Using existing sheet: {sanitize_excel_sheet_name(compressor_name)}")
                
                # Extract data to the appropriate worksheet
                success, page_count, rows_written, error_msg = extract_pdf_to_worksheet(
                    pdf_path, worksheet, compressor_name=compressor_name, show_progress=True
                )
                
                if success:
                    success_count += 1
                    total_rows += rows_written
                    compressor_stats[compressor_name] = compressor_stats.get(compressor_name, 0) + rows_written
                    print(f"  ✓ Success: {page_count} pages, {rows_written:,} rows")
                else:
                    failure_count += 1
                    failed_files.append((str(pdf_path), error_msg))
                    print(f"  ✗ Failed: {error_msg}")
                
            except Exception as e:
                failure_count += 1
                error_msg = str(e)
                failed_files.append((str(pdf_path), error_msg))
                print(f"  ✗ Error: {error_msg}")
            
            print()  # Blank line between files
        
        # Save Excel workbook
        print(f"Saving Excel file...")
        workbook.save(excel_output_path)
        workbook.close()
        
        # Get file size
        file_size_bytes = Path(excel_output_path).stat().st_size
        file_size_mb = file_size_bytes / (1024 * 1024)
        
        overall_time = time.time() - overall_start_time
        
        # Print summary
        print("=" * 60)
        print("PROCESSING SUMMARY")
        print("=" * 60)
        print(f"Total PDFs processed: {len(pdf_files)}")
        print(f"  Successful: {success_count}")
        print(f"  Failed: {failure_count}")
        print(f"Total rows written: {total_rows:,}")
        print(f"Total sheets created: {len(worksheets)}")
        print(f"Total processing time: {overall_time:.2f} seconds")
        print(f"Output file size: {file_size_mb:.2f} MB")
        print(f"Output file: {excel_output_path}")
        
        if compressor_stats:
            print()
            print("Rows per compressor:")
            for compressor_name, row_count in sorted(compressor_stats.items()):
                sheet_name = sanitize_excel_sheet_name(compressor_name)
                print(f"  {sheet_name}: {row_count:,} rows")
        
        if failed_files:
            print()
            print("Failed files:")
            for failed_file, error_msg in failed_files:
                print(f"  - {Path(failed_file).name}: {error_msg}")
        
        print()
        
        return {
            'success_count': success_count,
            'failure_count': failure_count,
            'total_rows': total_rows,
            'failed_files': failed_files,
            'total_time': overall_time,
            'compressor_stats': compressor_stats,
            'sheet_count': len(worksheets)
        }
        
    except Exception as e:
        print(f"✗ Fatal error: {e}")
        import traceback
        traceback.print_exc()
        # Try to save workbook if it exists
        if workbook is not None:
            try:
                workbook.save(excel_output_path)
                workbook.close()
                print(f"✓ Saved partial Excel file: {excel_output_path}")
            except:
                pass
        raise


def extract_with_pdfminer(pdf_path, output_path):
    """Extract text using pdfminer.six library."""
    from pdfminer.pdfpage import PDFPage
    from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
    from pdfminer.converter import TextConverter
    from pdfminer.layout import LAParams
    from io import StringIO
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Extracting with pdfminer.six...")
    print(f"  Opening PDF file...")
    start_time = time.time()
    open_start = time.time()
    
    try:
        BATCH_SIZE = 100
        
        # Count pages first
        with open(pdf_path, 'rb') as pdf_file:
            page_count = len(list(PDFPage.get_pages(pdf_file)))
        
        open_time = time.time() - open_start
        print(f"  ✓ PDF opened in {open_time:.2f} seconds")
        print(f"  Found {page_count} pages")
        print(f"  Starting page extraction (batching {BATCH_SIZE} pages at a time to reduce memory usage)...")
        
        extract_start = time.time()
        last_log_time = extract_start
        
        # Incremental statistics instead of storing all page_times
        total_time = 0.0
        min_time = float('inf')
        max_time = 0.0
        slowest_page_num = 0
        
        batch_buffer = []
        file_size_bytes = 0  # Track file size manually
        
        # Open file once and keep it open between batches
        output_file = open(output_path, 'w', encoding='utf-8')
        
        with open(pdf_path, 'rb') as pdf_file:
            pages = PDFPage.get_pages(pdf_file)
            
            for i, page in enumerate(pages):
                page_start = time.time()
                
                # Extract text from this page
                resource_manager = PDFResourceManager()
                output_string = StringIO()
                laparams = LAParams()
                device = TextConverter(resource_manager, output_string, laparams=laparams)
                interpreter = PDFPageInterpreter(resource_manager, device)
                interpreter.process_page(page)
                page_text = output_string.getvalue()
                device.close()
                output_string.close()
                
                page_time = time.time() - page_start
                
                # Update incremental statistics
                total_time += page_time
                if page_time < min_time:
                    min_time = page_time
                if page_time > max_time:
                    max_time = page_time
                    slowest_page_num = i + 1
                
                # Add page to batch buffer
                if page_text:
                    batch_buffer.append(f"=== PAGE {i + 1} ===\n{page_text}\n")
                
                # Write batch every BATCH_SIZE pages
                if len(batch_buffer) >= BATCH_SIZE:
                    batch_text = ''.join(batch_buffer)
                    output_file.write(batch_text)
                    output_file.flush()
                    file_size_bytes += len(batch_text.encode('utf-8'))
                    batch_buffer = []  # Clear buffer to free memory
                
                # Log progress every 10 pages or every 5 seconds
                current_time = time.time()
                if (i + 1) % 10 == 0 or (current_time - last_log_time) >= 5:
                    elapsed = current_time - extract_start
                    avg_time_per_page = elapsed / (i + 1)
                    remaining_pages = page_count - (i + 1)
                    est_remaining = avg_time_per_page * remaining_pages
                    file_size_mb = file_size_bytes / (1024 * 1024)
                    batch_num = (i + 1) // BATCH_SIZE + (1 if (i + 1) % BATCH_SIZE > 0 else 0)
                    print(f"  [{datetime.now().strftime('%H:%M:%S')}] Page {i + 1}/{page_count} "
                          f"(batch {batch_num}, "
                          f"avg: {avg_time_per_page:.3f}s/page, "
                          f"last: {page_time:.3f}s, "
                          f"file: {file_size_mb:.2f} MB, "
                          f"est. remaining: {est_remaining/60:.1f} min)")
                    last_log_time = current_time
                    sys.stdout.flush()
        
        # Write remaining batch
        if batch_buffer:
            batch_text = ''.join(batch_buffer)
            output_file.write(batch_text)
            file_size_bytes += len(batch_text.encode('utf-8'))
        
        output_file.close()
        
        extract_time = time.time() - extract_start
        avg_time = total_time / page_count if page_count > 0 else 0
        print(f"  ✓ Page extraction completed in {extract_time:.2f} seconds")
        print(f"  Average time per page: {avg_time:.3f} seconds")
        print(f"  Fastest page: {min_time:.3f} seconds")
        print(f"  Slowest page: {max_time:.3f} seconds (page {slowest_page_num})")
        
        # Read file to get character count
        print(f"  Reading file for statistics...")
        with open(output_path, 'r', encoding='utf-8') as f:
            full_text = f.read()
        
    except Exception as e:
        print(f"  ✗ Error with pdfminer.six: {e}")
        import traceback
        traceback.print_exc()
        if 'output_file' in locals():
            try:
                output_file.close()
            except:
                pass
        return None, None, None
    
    elapsed_time = time.time() - start_time
    
    char_count = len(full_text)
    file_size_mb = file_size_bytes / (1024 * 1024)
    print(f"  ✓ File complete: {file_size_mb:.2f} MB, {char_count:,} characters")
    
    return page_count, char_count, elapsed_time


def main():
    """Main function to run PDF extraction and write to Excel."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Extract compressor data from PDF(s) and write to Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Single PDF file
  python extract_pdf_text.py input.pdf
  python extract_pdf_text.py input.pdf -o output.xlsx
  
  # Multiple PDF files
  python extract_pdf_text.py file1.pdf file2.pdf file3.pdf -o combined.xlsx
  
  # Process all PDFs in a directory
  python extract_pdf_text.py -d tests/data/ -o combined.xlsx
        """
    )
    
    parser.add_argument(
        'pdf_files',
        type=str,
        nargs='*',
        help='Path(s) to input PDF file(s)'
    )
    
    parser.add_argument(
        '-d', '--directory',
        type=str,
        default=None,
        help='Process all PDF files in the specified directory'
    )
    
    parser.add_argument(
        '-o', '--output',
        type=str,
        default=None,
        help='Path to output Excel file (required for multiple PDFs, default: PDF filename with .xlsx extension for single file)'
    )
    
    args = parser.parse_args()
    
    # Determine mode: directory, multiple files, or single file
    pdf_files_to_process = []
    
    if args.directory:
        # Directory mode: find all PDFs in directory
        try:
            pdf_files_to_process = find_pdf_files(args.directory)
            if not pdf_files_to_process:
                print(f"Error: No PDF files found in directory: {args.directory}")
                return
        except Exception as e:
            print(f"Error: {e}")
            return
    elif args.pdf_files:
        # Multiple files or single file mode
        pdf_files_to_process = [Path(f) for f in args.pdf_files]
        
        # Validate all files exist
        for pdf_file in pdf_files_to_process:
            if not pdf_file.exists():
                print(f"Error: PDF file not found: {pdf_file}")
                return
    else:
        # Default: backward compatibility - use default file
        script_dir = Path(__file__).parent
        default_pdf = script_dir / "tests" / "data" / "COMPRESOR4-ABR-JUN-25.PDF"
        if default_pdf.exists():
            pdf_files_to_process = [default_pdf]
        else:
            print("Error: No PDF files specified and default file not found.")
            print("Usage: python extract_pdf_text.py <pdf_file> [pdf_file2 ...] [-o output.xlsx]")
            print("   or: python extract_pdf_text.py -d <directory> [-o output.xlsx]")
            return
    
    # Determine output Excel path
    if len(pdf_files_to_process) > 1:
        # Multiple files: output is required
        if not args.output:
            print("Error: Output file (-o/--output) is required when processing multiple PDFs")
            return
        excel_output_path = Path(args.output)
    else:
        # Single file: use provided output or default
        if args.output:
            excel_output_path = Path(args.output)
        else:
            excel_output_path = pdf_files_to_process[0].with_suffix('.xlsx')
    
    # Process files
    if len(pdf_files_to_process) == 1:
        # Single file mode (backward compatible)
        pdf_path = pdf_files_to_process[0]
        
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Starting PDF extraction to Excel")
        print(f"PDF file: {pdf_path}")
        print(f"PDF size: {pdf_path.stat().st_size / (1024*1024):.2f} MB")
        print(f"Excel output: {excel_output_path}")
        print(f"Active library: PyMuPDF\n")
        
        # Extract with PyMuPDF and write to Excel
        print("=" * 60)
        result = extract_with_pymupdf(pdf_path, excel_output_path)
        if result and result[0]:
            page_count, row_count, elapsed = result
            print(f"  ✓ Completed: {page_count} pages, "
                  f"{row_count:,} data rows, "
                  f"{elapsed:.2f} seconds")
            print(f"  Saved to: {excel_output_path}")
        print()
        
        # Summary
        print("=" * 60)
        print("EXTRACTION SUMMARY")
        print("=" * 60)
        if result and result[0]:
            page_count, row_count, elapsed = result
            print(f"Pages processed: {page_count}")
            print(f"Data rows written: {row_count:,}")
            print(f"Total time: {elapsed:.2f} seconds")
            print(f"Output file: {excel_output_path}")
        print()
    else:
        # Multiple files mode
        try:
            summary = process_multiple_pdfs(pdf_files_to_process, excel_output_path)
            # Summary is already printed by process_multiple_pdfs()
        except Exception as e:
            print(f"Error processing multiple PDFs: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)


if __name__ == "__main__":
    main()

